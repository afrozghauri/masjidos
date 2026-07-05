"""Acquisition MCP server.

Exposes tools the agent uses to GET a timetable from a masjid URL. The agent
decides which of these to call based on what it observes — it is not a fixed
pipeline. Each tool returns text or a base64 image the Comprehension server can read.
"""
import base64
import io
from pathlib import Path
import re
import httpx
import pandas as pd
import pdfplumber
from bs4 import BeautifulSoup
from mcp.server.fastmcp import FastMCP

import sys
sys.path.append(str(Path(__file__).resolve().parent.parent))
from config.settings import settings  # noqa: E402

mcp = FastMCP("acquisition")

HEADERS = {"User-Agent": "Mozilla/5.0 (MasjidOS timetable agent)"}


def _visible_text(soup: BeautifulSoup) -> str:
    """Plain visible body text, stripped of script/style noise. Jumuah times are
    often announced as a text banner rather than inside a <table>, so callers
    that only look at table_text can miss them."""
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(" ", strip=True)


def _open_google_worksheet(url: str):
    """Open the target worksheet via the SAME service account used for writing.
    The sheet only needs Editor access for the service account's client_email —
    it no longer needs to be public ('Anyone with the link can view')."""
    import gspread
    from google.oauth2.service_account import Credentials

    if not settings.google_service_account_json:
        raise ValueError("GOOGLE_SERVICE_ACCOUNT_JSON not set in .env — required to "
                         "read AND write a Google Sheet. Share the sheet with the "
                         "service account's client_email as Editor.")
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not m:
        raise ValueError("Could not parse a spreadsheet ID from that Google Sheets URL.")
    sheet_id = m.group(1)
    gid_match = re.search(r"[?&#]gid=(\d+)", url)
    gid = int(gid_match.group(1)) if gid_match else 0

    creds = Credentials.from_service_account_file(
        settings.google_service_account_json,
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    return next((w for w in sh.worksheets() if w.id == gid), sh.sheet1)


def _load_directory_df(source: str):
    """Load the masjid directory from EITHER:
    - a Google Sheets link (service account must have Editor access — no public
      sharing needed), or
    - a local .xlsx/.xls/.csv file path.
    The agent doesn't need to know which one it got — this resolves it."""
    if source.strip().lower().startswith("http"):
        if "docs.google.com/spreadsheets" in source:
            ws = _open_google_worksheet(source)
            records = ws.get_all_records()
            return pd.DataFrame(records)
        raise ValueError("Only Google Sheets links are supported as URLs right now.")
    path = Path(source)
    if not path.exists():
        raise FileNotFoundError(f"directory not found at {path}")
    return pd.read_excel(path) if path.suffix in (".xlsx", ".xls") else pd.read_csv(path)


@mcp.tool()
def read_directory(source: str = "", limit: int = 0) -> list[dict]:
    """Read the input spreadsheet of masjid names + URLs. `source` may be:
    - left empty (falls back to settings.masjid_directory, e.g. a local file path),
    - a Google Sheets link (shared as 'Anyone with the link can view'), or
    - a local .xlsx/.xls/.csv file path.
    `limit` caps how many rows are returned (0 = no limit).
    Returns a list of {"name", "url", "row_number", "source"} — row_number and
    source let mark_row_done write "Done" back into the exact right row later.
    This is the ONLY data input."""
    src = source.strip() if source else settings.masjid_directory
    try:
        df = _load_directory_df(src)
    except Exception as e:
        return [{"error": str(e)}]

    df.columns = [c.strip().lower() for c in df.columns]
    name_col = next((c for c in df.columns if "name" in c), df.columns[0])
    url_col = next((c for c in df.columns if "url" in c or "link" in c or "web" in c), df.columns[-1])
    out = []
    for idx, r in df.iterrows():
        name, url = str(r[name_col]).strip(), str(r[url_col]).strip()
        if name and url and url.lower() != "nan":
            out.append({"name": name, "url": url,
                       "row_number": int(idx) + 2,   # +2: header row + 0-index offset
                       "source": src})
        if limit and len(out) >= limit:
            break
    return out


@mcp.tool()
def find_timetable_page(url: str) -> dict:
    """Given a masjid's homepage (not necessarily the timetable page), list the
    site's internal links with their anchor text so the agent can REASON about
    which one is most likely the prayer/Salah/Iqamah timetable page. Masjids label
    this page inconsistently ("Prayer Times", "Salah Timetable", "Namaz Schedule",
    "Athan", "Timings", etc.) so this cannot be a fixed keyword match — the agent
    should look at the returned (text, href) pairs and decide."""
    try:
        resp = httpx.get(url, headers=HEADERS, timeout=30, follow_redirects=True)
        resp.raise_for_status()
    except Exception as e:
        return {"ok": False, "error": str(e)}
    soup = BeautifulSoup(resp.text, "lxml")
    base = str(resp.url)
    candidates = []
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        text = a.get_text(" ", strip=True)
        if not href or href.startswith("#") or href.startswith("mailto:"):
            continue
        full = httpx.URL(base).join(href) if not href.lower().startswith("http") else httpx.URL(href)
        full = str(full)
        if full in seen or full == base:
            continue
        seen.add(full)
        candidates.append({"text": text[:80], "href": full})
    return {"ok": True, "homepage": base, "links": candidates[:60],
            "note": "Pick the link whose text most plausibly refers to a prayer "
                    "timetable page, then call fetch_html on it. If none looks "
                    "right, this masjid's timetable may be on the homepage itself "
                    "or embedded as an image/widget — check the original page's "
                    "image_srcs from fetch_html."}


@mcp.tool()
def fetch_html(url: str) -> dict:
    """Fetch a page and return (a) any HTML tables found as text, and (b) a
    hint about whether the timetable is likely an image/PDF instead."""
    try:
        resp = httpx.get(url, headers=HEADERS, timeout=30, follow_redirects=True)
        resp.raise_for_status()
    except Exception as e:
        return {"ok": False, "error": str(e)}
    soup = BeautifulSoup(resp.text, "lxml")
    tables = soup.find_all("table")
    table_text = "\n\n".join(t.get_text(" ", strip=True) for t in tables[:5])
    pdf_links = [a.get("href") for a in soup.find_all("a", href=True)
                 if a["href"].lower().endswith(".pdf")]
    imgs = [i.get("src") for i in soup.find_all("img") if i.get("src")]
    page_text = _visible_text(soup)
    return {
        "ok": True,
        "table_text": table_text[:24000],
        "has_tables": bool(tables),
        "pdf_links": pdf_links[:5],
        "image_srcs": imgs[:10],
        "page_text": page_text[:6000],
        "note": "If has_tables is False, consider find_pdf_link or render_image_from_url. "
                "page_text is the general visible text on the page (outside <table> "
                "tags too) — useful for a Jumuah/Friday-prayer time announced as a "
                "text banner rather than a table cell.",
    }


@mcp.tool()
async def fetch_rendered_html(url: str) -> dict:
    """Fetch a page with a headless browser (JavaScript executed) instead of a
    plain HTTP GET. Some masjid sites inject the timetable client-side via a
    widget/plugin (e.g. jQuery-based prayer-time plugins) after page load, so
    fetch_html sees an empty shell. Use this as a fallback when fetch_html and
    find_timetable_page both report no tables, no PDF links, and no plausible
    timetable image."""
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page(user_agent=HEADERS["User-Agent"])
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)  # let client-side widgets/plugins populate
            html = await page.content()
            await browser.close()
    except Exception as e:
        return {"ok": False, "error": str(e)}
    soup = BeautifulSoup(html, "lxml")
    tables = soup.find_all("table")
    table_text = "\n\n".join(t.get_text(" ", strip=True) for t in tables[:5])
    pdf_links = [a.get("href") for a in soup.find_all("a", href=True)
                 if a["href"].lower().endswith(".pdf")]
    imgs = [i.get("src") for i in soup.find_all("img") if i.get("src")]
    page_text = _visible_text(soup)
    return {
        "ok": True,
        "table_text": table_text[:24000],
        "has_tables": bool(tables),
        "pdf_links": pdf_links[:5],
        "image_srcs": imgs[:10],
        "page_text": page_text[:6000],
        "note": "Rendered with a headless browser (JS executed). If has_tables is "
                "still False, the timetable may be a canvas/image widget — check "
                "image_srcs. page_text is the general visible page text, useful for "
                "a Jumuah/Friday-prayer time announced as a text banner.",
    }


@mcp.tool()
def find_pdf_link(url: str) -> dict:
    """Return the first likely timetable PDF link found on the page."""
    try:
        resp = httpx.get(url, headers=HEADERS, timeout=30, follow_redirects=True)
        soup = BeautifulSoup(resp.text, "lxml")
        links = [a["href"] for a in soup.find_all("a", href=True)
                 if a["href"].lower().endswith(".pdf")]
        return {"ok": True, "pdf_links": links[:5]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@mcp.tool()
def pdf_to_text(pdf_url: str) -> dict:
    """Download a PDF and extract its text layer. Empty text => it is an image
    PDF and the agent should fall back to render_pdf_page_image instead."""
    try:
        data = httpx.get(pdf_url, headers=HEADERS, timeout=60, follow_redirects=True).content
        text = ""
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages[:3]:
                text += (page.extract_text() or "") + "\n"
        return {"ok": True, "text": text[:24000],
                "is_image_pdf": len(text.strip()) < 40}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@mcp.tool()
def render_pdf_page_image(pdf_url: str, page: int = 0) -> dict:
    """Render a PDF page to a base64 PNG so the VLM can read image-only PDFs."""
    try:
        import fitz  # pymupdf
        data = httpx.get(pdf_url, headers=HEADERS, timeout=60, follow_redirects=True).content
        doc = fitz.open(stream=data, filetype="pdf")
        pix = doc[page].get_pixmap(dpi=150)
        b64 = base64.b64encode(pix.tobytes("png")).decode()
        return {"ok": True, "image_base64": b64}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@mcp.tool()
def render_image_from_url(image_url: str) -> dict:
    """Download an image URL and return it as base64 for the VLM."""
    try:
        data = httpx.get(image_url, headers=HEADERS, timeout=60, follow_redirects=True).content
        return {"ok": True, "image_base64": base64.b64encode(data).decode()}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _mark_done_google_sheet(url: str, row_number: int) -> dict:
    try:
        ws = _open_google_worksheet(url)
    except Exception as e:
        return {"ok": False, "error": str(e)}
    header = ws.row_values(1)
    target_col = len(header) + 1
    ws.update_cell(row_number, target_col, "Done")
    return {"ok": True, "row": row_number, "col": target_col, "sheet": ws.title}


def _mark_done_local_file(path_str: str, row_number: int) -> dict:
    path = Path(path_str)
    if not path.exists():
        return {"ok": False, "error": f"file not found at {path}"}
    if path.suffix in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        last_col = ws.max_column
        while last_col > 1 and ws.cell(row=1, column=last_col).value in (None, ""):
            last_col -= 1
        ws.cell(row=row_number, column=last_col + 1, value="Done")
        wb.save(path)
        return {"ok": True, "row": row_number, "col": last_col + 1, "file": str(path)}
    else:
        df = pd.read_csv(path)
        if "Status" not in df.columns:
            df["Status"] = ""
        df.loc[row_number - 2, "Status"] = "Done"
        df.to_csv(path, index=False)
        return {"ok": True, "row": row_number, "file": str(path),
                "note": "CSV rewritten with a Status column (no per-cell write for CSV)."}


@mcp.tool()
def mark_row_done(source: str, row_number: int) -> dict:
    """Write 'Done' in the cell immediately to the right of the last column of the
    given row, in the SAME spreadsheet the directory was read from — Google Sheet
    (needs a service account with Editor access) or local Excel/CSV file.
    row_number is the 1-indexed row as returned by read_directory."""
    src = source.strip()
    try:
        if src.lower().startswith("http") and "docs.google.com/spreadsheets" in src:
            return _mark_done_google_sheet(src, row_number)
        return _mark_done_local_file(src, row_number)
    except Exception as e:
        return {"ok": False, "error": str(e)}

if __name__ == "__main__":
    mcp.run()
